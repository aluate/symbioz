"""
Invoice generation module.
Creates Excel invoice matching the Moulding Calculator Template format.
"""

import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def load_bf_cost():
    """Load BF cost per species and thickness from Excel"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, 'bf_cost.xlsx')
    
    if not os.path.exists(excel_path):
        return None
    
    try:
        df = pd.read_excel(excel_path)
        return df
    except Exception:
        return None


def get_bf_cost(species, thickness_category, bf_cost_df):
    """Get BF cost for a species and thickness category"""
    if bf_cost_df is None:
        return 0.0
    
    try:
        row = bf_cost_df[bf_cost_df['Species'] == species]
        if row.empty:
            return 0.0
        cost = row[thickness_category].iloc[0]
        return float(cost) if pd.notna(cost) else 0.0
    except Exception:
        return 0.0


def get_finish_multiplier(is_finishing, finish_type):
    """Get finish price multiplier (1.0 for no finish, 1.15 for finish)"""
    if not is_finishing:
        return 1.0
    # Finish adds 15% multiplier
    return 1.15


def get_setup_cost(trim_type, finish_level):
    """Get setup cost based on trim type and finish level (simplified - can be enhanced)"""
    # Base setup costs (can be made configurable)
    setup_costs = {
        'base': 725,
        'casing': 525,
        'headers': 725,
        'sills': 925,
        'apron': 725,
        'jambs': 675,  # INT JAMB
        'dentils': 800
    }
    return setup_costs.get(trim_type, 700)


def format_description(species, trim_type, width, thickness, finish_level, is_finishing, finish_type, linear_ft, bf_required):
    """Format description: Base – Paint – 5.5" – 128 LF – 42.37 BF – Clear Finish"""
    # Map trim types to display names
    type_map = {
        'base': 'Base',
        'casing': 'Casing',
        'headers': 'Headers',
        'sills': 'Sills',
        'apron': 'Apron',
        'jambs': 'Jambs',
        'dentils': 'Dentils'
    }
    type_name = type_map.get(trim_type, trim_type.title())
    
    # Format finish type
    if is_finishing and finish_type:
        finish_label = finish_type.title()
    else:
        finish_label = 'Primed'
    
    # Round values
    width_rounded = round(width, 1)
    thickness_rounded = round(thickness, 1)
    lf_rounded = round(linear_ft, 0)  # Round LF to whole number for description
    bf_rounded = round(bf_required, 2)
    
    # Format: "Base – Paint – 5.5" – 128 LF – 42.37 BF – Clear Finish"
    description = f"{type_name} – {finish_label} – {width_rounded}\" – {lf_rounded} LF – {bf_rounded} BF"
    
    # Add finish detail if finishing
    if is_finishing and finish_type:
        description += f" – {finish_type.title()} Finish"
    
    return description


def calculate_pricing(lf_results, bf_data, species, lf_cost_per_ft, bf_markup_pct, is_finishing, finish_type):
    """Calculate pricing for trim items"""
    bf_cost_df = load_bf_cost()
    pricing = {}
    
    for trim_type, linear_ft in lf_results.items():
        if linear_ft == 0:
            continue
        
        bf_required = bf_data['bf_with_waste'].get(trim_type, 0.0)
        thickness_category = bf_data['thickness_category'].get(trim_type, '')
        width, finished_thickness = bf_data['dimensions'].get(trim_type, (0.0, 0.0))
        
        # Get BF cost from Excel
        bf_cost_per_bf = get_bf_cost(species, thickness_category, bf_cost_df)
        
        # Calculate costs
        bf_cost = bf_required * bf_cost_per_bf
        bf_cost_with_markup = bf_cost * (1 + bf_markup_pct / 100)
        
        # Price per LF = (BF cost with markup / LF) + LF cost per foot
        if linear_ft > 0:
            base_price_per_lf = (bf_cost_with_markup / linear_ft) + lf_cost_per_ft
        else:
            base_price_per_lf = 0.0
        
        # Apply finish multiplier
        finish_mult = get_finish_multiplier(is_finishing, finish_type)
        price_per_lf = base_price_per_lf * finish_mult
        
        # Get setup cost
        finish_level = 'Standard'  # Could be passed in
        setup_cost = get_setup_cost(trim_type, finish_level)
        
        # Line total = (Price per LF * LF) + Setup
        line_total = (price_per_lf * linear_ft) + setup_cost
        
        pricing[trim_type] = {
            'linear_ft': linear_ft,
            'bf_required': bf_required,
            'bf_cost_per_bf': bf_cost_per_bf,
            'bf_cost_total': bf_cost,
            'bf_cost_with_markup': bf_cost_with_markup,
            'lf_cost_per_ft': lf_cost_per_ft,
            'price_per_lf': price_per_lf,
            'base_price_per_lf': base_price_per_lf,
            'finish_multiplier': finish_mult,
            'setup_cost': setup_cost,
            'line_total': line_total,
            'thickness_category': thickness_category,
            'width': width,
            'thickness': finished_thickness
        }
    
    return pricing


def setup_output_folders(script_dir):
    """Create Output and archive folders if they don't exist"""
    output_dir = os.path.join(script_dir, 'Output')
    archive_dir = os.path.join(output_dir, 'archive')
    
    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(archive_dir, exist_ok=True)
    
    return output_dir, archive_dir


def archive_old_invoices(output_dir, archive_dir):
    """Move existing invoice files from Output to archive"""
    import shutil
    
    # Find all Excel files in Output (excluding archive folder)
    for filename in os.listdir(output_dir):
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            source = os.path.join(output_dir, filename)
            destination = os.path.join(archive_dir, filename)
            
            # If file already exists in archive, add timestamp
            if os.path.exists(destination):
                name, ext = os.path.splitext(filename)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                destination = os.path.join(archive_dir, f"{name}_{timestamp}{ext}")
            
            try:
                shutil.move(source, destination)
            except Exception:
                pass  # Skip if move fails


def generate_invoice_excel(all_results, species, lf_cost_per_ft, bf_markup_pct, is_finishing, finish_type, 
                          job_name='', customer_name='', sales_tax_rate=0.06, output_path=None):
    """
    Generate Excel invoice matching Moulding Calculator Template format.
    
    Args:
        all_results: Complete calculation results
        species: Species name
        lf_cost_per_ft: Cost per linear foot
        bf_markup_pct: BF markup percentage
        is_finishing: Whether finishing is required
        finish_type: Finish type (Stain/Paint/Other)
        job_name: Job name
        customer_name: Customer name
        sales_tax_rate: Sales tax rate (default 0.06 = 6%)
        output_path: Path to save Excel file (if None, uses Output folder)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    
    # Setup output folders
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir, archive_dir = setup_output_folders(script_dir)
    
    # If no output path provided, generate one in Output folder
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        job_suffix = f"_{job_name.replace(' ', '_')}" if job_name else ""
        filename = f"Trim_Estimate_{timestamp}{job_suffix}.xlsx"
        output_path = os.path.join(output_dir, filename)
    
    # Archive old invoices before creating new one
    archive_old_invoices(output_dir, archive_dir)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Trim Estimate"
    
    # Import styles
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # Define styles
    header_font = Font(name='Arial', size=10, bold=True)
    header_alignment = Alignment(horizontal='right', vertical='center')
    title_font = Font(name='Calibri', size=14)
    date_font = Font(name='Calibri', size=11)
    label_font = Font(name='Arial', size=10, bold=True)
    number_alignment = Alignment(horizontal='right', vertical='center')
    text_alignment = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Row 1: Empty (logo placeholder)
    ws.merge_cells('C1:C1')
    ws['C1'] = ''  # Logo placeholder
    
    # Row 2: Estimate and Date
    ws['D2'] = 'Estimate'
    ws['D2'].font = title_font
    ws['G2'] = f"Date: {datetime.now().strftime('%Y-%m-%d')}"
    ws['G2'].font = date_font
    
    # Row 3: Job and Sales Tax
    ws['G3'] = f"Job: {job_name}"
    ws['G3'].font = label_font
    ws['G3'].alignment = Alignment(horizontal='right', vertical='center')
    ws['I3'] = 'Sales Tax'
    ws['I3'].font = label_font
    ws['I3'].alignment = header_alignment
    ws['J3'] = sales_tax_rate
    ws['J3'].number_format = '0.00%'
    
    # Row 4: Customer
    ws['G4'] = f"Customer: {customer_name}"
    ws['G4'].font = label_font
    ws['G4'].alignment = Alignment(horizontal='right', vertical='center')
    
    # Row 5: Header row
    headers = [
        ('B5', 'Line Number'),
        ('C5', 'Qty LF'),
        ('D5', 'Description'),
        ('E5', 'LF Price'),
        ('F5', 'Set Up'),
        ('G5', 'Finish'),
        ('H5', 'LF Price'),
        ('I5', 'Line Total')
    ]
    for cell_ref, header_text in headers:
        ws[cell_ref] = header_text
        ws[cell_ref].font = header_font
        ws[cell_ref].alignment = header_alignment
        ws[cell_ref].border = thin_border
    
    # Collect all line items across all styles
    line_items = []
    line_number = 1
    
    for style, data in all_results['styles'].items():
        lf_results = data['lf']
        bf_data = data['bf']
        finish_level = all_results['inputs'].get('finish_level', 'Standard')
        
        # Calculate pricing for this style
        pricing = calculate_pricing(lf_results, bf_data, species, lf_cost_per_ft, bf_markup_pct, 
                                  is_finishing, finish_type)
        
        trim_type_order = ['base', 'casing', 'headers', 'sills', 'apron', 'jambs', 'dentils']
        for trim_type in trim_type_order:
            if trim_type in pricing:
                p = pricing[trim_type]
                if p['linear_ft'] > 0:
                    description = format_description(
                        species, trim_type, p['width'], p['thickness'], 
                        finish_level, is_finishing, finish_type,
                        p['linear_ft'], p['bf_required']
                    )
                    
                    line_items.append({
                        'line_number': line_number,
                        'qty_lf': round(p['linear_ft'], 2),
                        'description': description,
                        'lf_price': round(p['base_price_per_lf'], 2),
                        'setup': round(p['setup_cost'], 2),
                        'finish': round(p['finish_multiplier'], 2),
                        'lf_price_final': round(p['price_per_lf'], 2),
                        'line_total': round(p['line_total'], 2)
                    })
                    line_number += 1
    
    # Write line items (rows 6-30)
    for idx, item in enumerate(line_items[:25]):  # Max 25 lines
        row = 6 + idx
        ws[f'B{row}'] = item['line_number']
        ws[f'B{row}'].alignment = number_alignment
        ws[f'C{row}'] = item['qty_lf']
        ws[f'C{row}'].number_format = '0.00'
        ws[f'C{row}'].alignment = number_alignment
        ws[f'D{row}'] = item['description']
        ws[f'D{row}'].alignment = text_alignment
        ws[f'E{row}'] = item['lf_price']
        ws[f'E{row}'].number_format = '$#,##0.00'
        ws[f'E{row}'].alignment = number_alignment
        ws[f'F{row}'] = item['setup']
        ws[f'F{row}'].number_format = '#,##0'
        ws[f'F{row}'].alignment = number_alignment
        ws[f'G{row}'] = item['finish']
        ws[f'G{row}'].number_format = '0.00'
        ws[f'G{row}'].alignment = number_alignment
        ws[f'H{row}'] = item['lf_price_final']
        ws[f'H{row}'].number_format = '$#,##0.00'
        ws[f'H{row}'].alignment = number_alignment
        ws[f'I{row}'] = item['line_total']
        ws[f'I{row}'].number_format = '$#,##0.00'
        ws[f'I{row}'].alignment = number_alignment
        
        # Add borders to data rows
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws[f'{col}{row}'].border = thin_border
    
    # Row 31: Species subtotal
    total_subtotal = round(sum(item['line_total'] for item in line_items), 2)
    ws['G31'] = species
    ws['G31'].font = label_font
    ws['G31'].alignment = Alignment(horizontal='right', vertical='center')
    ws['I31'] = total_subtotal
    ws['I31'].number_format = '$#,##0.00'
    ws['I31'].font = Font(bold=True)
    ws['I31'].alignment = number_alignment
    ws['I31'].border = thin_border
    
    # Row 34: Total
    tax_amount = round(total_subtotal * sales_tax_rate, 2)
    total_with_tax = round(total_subtotal + tax_amount, 2)
    ws['G34'] = 'Total'
    ws['G34'].font = Font(name='Arial', size=10, bold=True)
    ws['G34'].alignment = Alignment(horizontal='right', vertical='center')
    ws['I34'] = total_with_tax
    ws['I34'].number_format = '$#,##0.00'
    ws['I34'].font = Font(name='Arial', size=10, bold=True)
    ws['I34'].alignment = number_alignment
    ws['I34'].border = thin_border
    
    # Set column widths for better readability
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    
    # Save file
    wb.save(output_path)
