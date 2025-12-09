"""Excel export for BOM."""

from pathlib import Path
from typing import List, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from .model import PartQuantity


# Category to sheet name mapping
CATEGORY_SHEETS = {
    "Framing": "Framing",
    "Doors_Windows": "Doors_Windows",
    "Metal": "Metal",
    "Insulation": "Insulation",
    "Concrete": "Concrete",
    "MEP": "MEP",
    "Misc": "Misc",
}


def export_bom_to_excel(
    bom: List[PartQuantity],
    output_path: Path,
    project_name: Optional[str] = None,
) -> None:
    """
    Export BOM to Excel with multiple category tabs.
    
    Args:
        bom: List of PartQuantity items
        output_path: Path to save Excel file
        project_name: Optional project name for filename
    """
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Group by export_category
    by_category: dict[str, List[PartQuantity]] = {}
    for item in bom:
        category = item.export_category or "Misc"
        if category not in by_category:
            by_category[category] = []
        by_category[category].append(item)
    
    # Create sheets for each category
    for category, items in by_category.items():
        sheet_name = CATEGORY_SHEETS.get(category, "Misc")
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
        else:
            ws = wb[sheet_name]
        
        # Headers
        headers = ["Part ID", "Part Name", "Description", "Length (in)", "Unit", "Qty", "Unit Price", "Ext Price", "Notes"]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        # Sort by part_name, then by length_in (None last)
        for item in sorted(items, key=lambda x: (x.part_name, x.length_in if x.length_in is not None else float('inf'))):
            row = [
                item.part_id,
                item.part_name,
                item.notes or "",
                item.length_in if item.length_in is not None else "",
                item.unit,
                item.qty,
                item.unit_price,
                item.ext_price,
                item.notes or "",
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create Summary sheet
    if "Summary" not in wb.sheetnames:
        summary_ws = wb.create_sheet(title="Summary", index=0)
    else:
        summary_ws = wb["Summary"]
    
    summary_ws.append(["Category", "Total Cost"])
    summary_ws[1][0].font = Font(bold=True)
    summary_ws[1][1].font = Font(bold=True)
    
    # Calculate totals by category
    category_totals: dict[str, float] = {}
    for item in bom:
        category = item.export_category or "Misc"
        category_totals[category] = category_totals.get(category, 0.0) + item.ext_price
    
    for category, total in sorted(category_totals.items()):
        summary_ws.append([category, total])
    
    # Grand total
    grand_total = sum(item.ext_price for item in bom)
    summary_ws.append(["", ""])
    summary_ws.append(["GRAND TOTAL", grand_total])
    summary_ws[summary_ws.max_row][0].font = Font(bold=True)
    summary_ws[summary_ws.max_row][1].font = Font(bold=True)
    
    # Auto-adjust summary column widths
    for column in summary_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        summary_ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(output_path)


def generate_bom_filename(project_name: Optional[str] = None) -> str:
    """Generate filename for BOM export."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if project_name:
        # Sanitize project name for filename
        safe_name = "".join(c for c in project_name if c.isalnum() or c in (" ", "-", "_")).strip()
        safe_name = safe_name.replace(" ", "_")
        return f"material_list_{safe_name}_{timestamp}.xlsx"
    return f"material_list_{timestamp}.xlsx"

